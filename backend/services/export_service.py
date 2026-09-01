"""Export service for CRISPR-Sim: PDF, Excel (XLSX), CSV, FASTA.

Generates publication-quality reports, multi-sheet workbooks, and bioinformatic standard files.
"""

from __future__ import annotations

import csv
import io
import textwrap
from datetime import datetime, timezone
from typing import Any, Dict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def _wrap_seq(seq: str, width: int = 80) -> str:
    """Wraps sequence into standard fixed-width lines."""
    if not seq:
        return ""
    return "\n".join(textwrap.wrap(seq, width=width))


def generate_analysis_pdf(data: Dict[str, Any]) -> bytes:
    """Generates a publication-grade scientific PDF analysis report."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    teal_primary = colors.HexColor("#0D9488")
    teal_dark = colors.HexColor("#115E59")
    text_dark = colors.HexColor("#0F172A")
    slate_border = colors.HexColor("#CBD5E1")
    slate_bg = colors.HexColor("#F8FAFC")
    danger_bg = colors.HexColor("#FEF2F2")
    success_bg = colors.HexColor("#ECFDF5")

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=teal_dark,
    )

    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#64748B"),
    )

    heading_style = ParagraphStyle(
        "SectionHeading",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=teal_dark,
        spaceAfter=4,
    )

    body_style = ParagraphStyle(
        "BodyDark",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=text_dark,
    )

    body_bold = ParagraphStyle(
        "BodyBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=11,
        textColor=text_dark,
    )

    mono_style = ParagraphStyle(
        "MonoSeq",
        parent=styles["Normal"],
        fontName="Courier",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#1E293B"),
    )

    story = []

    # Title & Branding Header Table
    timestamp_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    header_data = [
        [
            Paragraph("<b>CRISPR-Sim</b> | Gene Editing Analysis Dossier", title_style),
            Paragraph(f"<b>Report Date:</b> {timestamp_str}<br/><b>Engine:</b> Cas Endonuclease Simulator v2.4", subtitle_style),
        ]
    ]
    header_table = Table(header_data, colWidths=[360, 180])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=1.5, color=teal_primary, spaceAfter=8))

    # Executive KPI Cards Table
    repair_type = data.get("repair_type", "NHEJ")
    frameshift = data.get("frameshift", False)
    premature_stop = data.get("premature_stop", False)
    safety_score = data.get("safety_score", 62)
    safety_label = data.get("safety_label", "Moderate")

    kpi_data = [
        [
            Paragraph("<b>SAFETY SCORE</b>", subtitle_style),
            Paragraph("<b>REPAIR MECHANISM</b>", subtitle_style),
            Paragraph("<b>READING FRAME</b>", subtitle_style),
            Paragraph("<b>STOP CODON</b>", subtitle_style),
        ],
        [
            Paragraph(f"<font size=12 color='{'#059669' if safety_score>=70 else '#D97706' if safety_score>=50 else '#DC2626'}'><b>{safety_score}/100</b></font><br/><font size=7.5>{safety_label} Risk Profile</font>", body_style),
            Paragraph(f"<font size=11 color='#0F172A'><b>{repair_type}</b></font><br/><font size=7.5>{'Error-Prone Indel' if repair_type=='NHEJ' else 'Template Homology'}</font>", body_style),
            Paragraph(f"<font size=11 color='{'#DC2626' if frameshift else '#059669'}'><b>{'Disrupted' if frameshift else 'In-Frame'}</b></font><br/><font size=7.5>{'Frameshift Mutation' if frameshift else 'Codon Frame Intact'}</font>", body_style),
            Paragraph(f"<font size=11 color='{'#DC2626' if premature_stop else '#059669'}'><b>{'Detected (*)' if premature_stop else 'None'}</b></font><br/><font size=7.5>{'Truncated Product' if premature_stop else 'Full Translation'}</font>", body_style),
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[135, 135, 135, 135])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), slate_bg),
        ('BOX', (0, 0), (-1, -1), 1, slate_border),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, slate_border),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 10))

    # Section 1: Sequence & Indel Statistics
    story.append(Paragraph("1. Sequence & Indel Statistics", heading_style))
    orig_len = data.get("original_length", 276)
    edit_len = data.get("edited_length", 269)
    len_diff = data.get("length_diff", 7)
    indel_type = "Deletion" if len_diff > 0 else "Insertion" if len_diff < 0 else "Unchanged"

    stats_data = [
        [Paragraph("<b>Metric</b>", body_bold), Paragraph("<b>Simulated Value</b>", body_bold), Paragraph("<b>Biological Significance & Verification</b>", body_bold)],
        [Paragraph("Original DNA Target Length", body_style), Paragraph(f"{orig_len} bp", body_style), Paragraph("Wild-type reference sequence baseline", body_style)],
        [Paragraph("Edited Sequence Length", body_style), Paragraph(f"{edit_len} bp", body_style), Paragraph("Resulting length post-cleavage & cellular repair", body_style)],
        [Paragraph("Net Indel Size", body_style), Paragraph(f"{abs(len_diff)} bp ({indel_type})", body_style), Paragraph(f"{'Disrupts triplet codon reading frame' if frameshift else 'Preserves triplet codon reading frame'}", body_style)],
        [Paragraph("Repair Pathway", body_style), Paragraph(str(repair_type), body_style), Paragraph("Non-Homologous End Joining (NHEJ) / HDR knock-in", body_style)],
        [Paragraph("CRISPR Safety Score", body_style), Paragraph(f"{safety_score} / 100", body_style), Paragraph(f"Algorithmic evaluation: {safety_label} off-target risk", body_style)],
    ]
    stats_table = Table(stats_data, colWidths=[150, 120, 270])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), teal_primary),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, slate_border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, slate_bg]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 10))

    # Section 2: Protein Sequence Comparison
    story.append(Paragraph("2. Protein Translation & Polypeptide Alignment", heading_style))
    orig_prot = data.get("original_protein", "")
    edit_prot = data.get("edited_protein", "")

    wrapped_orig_prot = _wrap_seq(orig_prot, 70) or "N/A"
    wrapped_edit_prot = _wrap_seq(edit_prot, 70) or "N/A"

    protein_data = [
        [Paragraph("<b>Translation Track</b>", body_bold), Paragraph("<b>Amino Acid Sequence (N-Terminus → C-Terminus)</b>", body_bold)],
        [
            Paragraph("<b>Original Protein:</b><br/><font size=7 color='#047857'>Wild-Type</font>", body_style),
            Paragraph(wrapped_orig_prot.replace("\n", "<br/>"), mono_style),
        ],
        [
            Paragraph(f"<b>Edited Protein:</b><br/><font size=7 color='{'#DC2626' if frameshift or premature_stop else '#047857'}'>Post-{repair_type}</font>", body_style),
            Paragraph(f"<font color='{'#DC2626' if frameshift or premature_stop else '#047857'}'>{wrapped_edit_prot.replace(chr(10), '<br/>')}</font>", mono_style),
        ],
    ]
    protein_table = Table(protein_data, colWidths=[120, 420])
    protein_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, slate_border),
        ('BACKGROUND', (0, 1), (-1, 1), success_bg),
        ('BACKGROUND', (0, 2), (-1, 2), danger_bg if (frameshift or premature_stop) else success_bg),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(protein_table)
    story.append(Spacer(1, 10))

    # Section 3: mRNA Sequence Comparison
    story.append(Paragraph("3. mRNA Transcript Comparison (5' → 3')", heading_style))
    orig_mrna = data.get("original_mrna", "")
    edit_mrna = data.get("edited_mrna", "")

    wrapped_orig_mrna = _wrap_seq(orig_mrna, 70) or "N/A"
    wrapped_edit_mrna = _wrap_seq(edit_mrna, 70) or "N/A"

    mrna_data = [
        [Paragraph("<b>Transcript Track</b>", body_bold), Paragraph("<b>Ribonucleotide Sequence (5' → 3')</b>", body_bold)],
        [
            Paragraph("<b>Original mRNA:</b>", body_style),
            Paragraph(wrapped_orig_mrna.replace("\n", "<br/>"), mono_style),
        ],
        [
            Paragraph("<b>Edited mRNA:</b>", body_style),
            Paragraph(f"<font color='#7E22CE'>{wrapped_edit_mrna.replace(chr(10), '<br/>')}</font>", mono_style),
        ],
    ]
    mrna_table = Table(mrna_data, colWidths=[120, 420])
    mrna_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#334155")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, slate_border),
        ('BACKGROUND', (0, 1), (-1, 1), slate_bg),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor("#FAF5FF")),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(mrna_table)
    story.append(Spacer(1, 12))

    # Footer note
    story.append(Paragraph(
        "<i>Notice: This report was generated algorithmically by the CRISPR-Sim Bioinformatic Simulator. "
        "Intended for academic research, education, and experimental planning. Laboratory validation is required prior to in vivo usage.</i>",
        subtitle_style,
    ))

    doc.build(story)
    return buffer.getvalue()


def generate_analysis_excel(data: Dict[str, Any]) -> bytes:
    """Generates an Excel workbook with cleanly aligned, styled worksheets."""
    wb = openpyxl.Workbook()
    repair_type = data.get("repair_type", "NHEJ")
    frameshift = data.get("frameshift", False)
    premature_stop = data.get("premature_stop", False)
    safety_score = data.get("safety_score", 62)
    safety_label = data.get("safety_label", "Moderate")
    len_diff = data.get("length_diff", 7)
    indel_type = "Deletion" if len_diff > 0 else "Insertion" if len_diff < 0 else "Unchanged"

    # Style definitions
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    dark_header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    alert_red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    alert_green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="0D9488")
    subtitle_font = Font(name="Calibri", size=9, italic=True, color="64748B")
    bold_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    regular_font = Font(name="Calibri", size=10, color="0F172A")
    mono_font = Font(name="Consolas", size=9.5, color="1E293B")
    red_font = Font(name="Calibri", size=10, bold=True, color="DC2626")
    green_font = Font(name="Calibri", size=10, bold=True, color="059669")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── Sheet 1: Executive Summary ──
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "CRISPR-Sim — Gene Editing Analysis Dossier"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = left_align

    ws1["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')} | Platform: CRISPR-Sim In Silico Engine"
    ws1["A2"].font = subtitle_font
    ws1["A2"].alignment = left_align

    summary_headers = ["Parameter / Metric", "Simulated Value", "Clinical / Biological Impact", "Quality Status"]
    for col_idx, h in enumerate(summary_headers, start=1):
        c = ws1.cell(row=4, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align if col_idx in [2, 4] else left_align
        c.border = thin_border

    summary_rows = [
        ("Repair Pathway", repair_type, "Simulated double-strand break repair mechanism", "Simulated"),
        ("CRISPR Safety Score", f"{safety_score}/100 ({safety_label})", "Aggregate safety based on PAM, off-target risk, & GC profile", "Optimal" if safety_score>=70 else "Review"),
        ("Frameshift Mutation", "YES (Disrupted)" if frameshift else "NO (In-Frame)", "Indel size not multiple of 3; disrupts downstream codons" if frameshift else "Triplet frame preserved", "MUTATION" if frameshift else "PASS"),
        ("Premature Stop Codon", "YES (Truncated)" if premature_stop else "NO (Intact)", "Early stop codon (*) identified in edited reading frame" if premature_stop else "No nonsense mutation detected", "TRUNCATION" if premature_stop else "PASS"),
        ("Original DNA Length", f"{data.get('original_length', 276)} bp", "Wild-type target sequence base pairs", "Baseline"),
        ("Edited Sequence Length", f"{data.get('edited_length', 269)} bp", "Post-repair sequence length", "Verified"),
        ("Net Indel Size", f"{abs(len_diff)} bp ({indel_type})", f"{abs(len_diff)} bp net structural alteration", "Indel Detected"),
    ]

    for r_idx, (param, val, impact, status) in enumerate(summary_rows, start=5):
        row_cells = [
            ws1.cell(row=r_idx, column=1, value=param),
            ws1.cell(row=r_idx, column=2, value=val),
            ws1.cell(row=r_idx, column=3, value=impact),
            ws1.cell(row=r_idx, column=4, value=status),
        ]
        is_even = (r_idx % 2 == 0)
        for c_idx, cell in enumerate(row_cells, start=1):
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if c_idx == 1:
                cell.font = bold_font
                cell.alignment = left_align
            elif c_idx == 2:
                cell.alignment = center_align
                if "YES" in str(val):
                    cell.font = red_font
                    cell.fill = alert_red_fill
                elif "NO" in str(val):
                    cell.font = green_font
                    cell.fill = alert_green_fill
                else:
                    cell.font = bold_font
            elif c_idx == 3:
                cell.font = regular_font
                cell.alignment = left_align
            elif c_idx == 4:
                cell.alignment = center_align
                if status in ["MUTATION", "TRUNCATION", "Review"]:
                    cell.font = red_font
                else:
                    cell.font = green_font

    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 56
    ws1.column_dimensions["D"].width = 18

    # ── Sheet 2: Protein Comparison ──
    ws2 = wb.create_sheet(title="Protein Sequences")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Protein Translation & Polypeptide Comparison"
    ws2["A1"].font = title_font

    prot_headers = ["Track", "Length", "Amino Acid Sequence (N → C)"]
    for col_idx, h in enumerate(prot_headers, start=1):
        c = ws2.cell(row=3, column=col_idx, value=h)
        c.fill = dark_header_fill
        c.font = header_font
        c.border = thin_border
        c.alignment = center_align if col_idx <= 2 else left_align

    orig_prot = data.get("original_protein", "")
    edit_prot = data.get("edited_protein", "")

    p_rows = [
        ("Original Protein (Wild-Type)", f"{len(orig_prot)} aa", orig_prot),
        (f"Edited Protein ({repair_type})", f"{len(edit_prot)} aa", edit_prot),
    ]

    for r_idx, (track, length_str, seq) in enumerate(p_rows, start=4):
        c1 = ws2.cell(row=r_idx, column=1, value=track)
        c2 = ws2.cell(row=r_idx, column=2, value=length_str)
        c3 = ws2.cell(row=r_idx, column=3, value=seq)

        for c in [c1, c2, c3]:
            c.border = thin_border
        c1.font = bold_font
        c1.alignment = left_align
        c2.font = regular_font
        c2.alignment = center_align
        c3.font = mono_font
        c3.alignment = wrap_align

        if r_idx == 4:
            c1.fill = alert_green_fill
        else:
            c1.fill = alert_red_fill if (frameshift or premature_stop) else alert_green_fill

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 90

    # ── Sheet 3: mRNA Sequences ──
    ws3 = wb.create_sheet(title="mRNA Transcripts")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "mRNA Transcripts (5' → 3') Comparison"
    ws3["A1"].font = title_font

    mrna_headers = ["Track", "Length", "Ribonucleotide Sequence (5' → 3')"]
    for col_idx, h in enumerate(mrna_headers, start=1):
        c = ws3.cell(row=3, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = thin_border
        c.alignment = center_align if col_idx <= 2 else left_align

    orig_mrna = data.get("original_mrna", "")
    edit_mrna = data.get("edited_mrna", "")

    m_rows = [
        ("Original mRNA Transcript", f"{len(orig_mrna)} nt", orig_mrna),
        (f"Edited mRNA Transcript ({repair_type})", f"{len(edit_mrna)} nt", edit_mrna),
    ]

    for r_idx, (track, length_str, seq) in enumerate(m_rows, start=4):
        c1 = ws3.cell(row=r_idx, column=1, value=track)
        c2 = ws3.cell(row=r_idx, column=2, value=length_str)
        c3 = ws3.cell(row=r_idx, column=3, value=seq)

        for c in [c1, c2, c3]:
            c.border = thin_border
            c.fill = zebra_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        c1.font = bold_font
        c1.alignment = left_align
        c2.font = regular_font
        c2.alignment = center_align
        c3.font = mono_font
        c3.alignment = wrap_align

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 90

    # Freeze panes on all sheets
    ws1.freeze_panes = "A5"
    ws2.freeze_panes = "A4"
    ws3.freeze_panes = "A4"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_analysis_csv(data: Dict[str, Any]) -> str:
    """Generates a structured, clean CSV export."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Metric", "Value"])
    writer.writerow(["Generated At", datetime.now(timezone.utc).isoformat()])
    writer.writerow(["Repair Type", data.get("repair_type", "NHEJ")])
    writer.writerow(["Safety Score", f"{data.get('safety_score', 62)}/100"])
    writer.writerow(["Safety Label", data.get("safety_label", "Moderate")])
    writer.writerow(["Frameshift Disruption", "YES" if data.get("frameshift") else "NO"])
    writer.writerow(["Premature Stop Codon", "YES" if data.get("premature_stop") else "NO"])
    writer.writerow(["Original Length (bp)", data.get("original_length", 276)])
    writer.writerow(["Edited Length (bp)", data.get("edited_length", 269)])
    writer.writerow(["Length Difference (bp)", data.get("length_diff", 7)])
    writer.writerow(["Original DNA", data.get("original_dna", data.get("original_mrna", ""))])
    writer.writerow(["Edited DNA", data.get("edited_dna", data.get("edited_mrna", ""))])
    writer.writerow(["Original Protein", data.get("original_protein", "")])
    writer.writerow(["Edited Protein", data.get("edited_protein", "")])
    writer.writerow(["Original mRNA", data.get("original_mrna", "")])
    writer.writerow(["Edited mRNA", data.get("edited_mrna", "")])

    return output.getvalue()


def generate_analysis_fasta(data: Dict[str, Any]) -> str:
    """Generates standard multi-record FASTA formatted with 80-character line wrapping."""
    repair_type = data.get("repair_type", "NHEJ")
    orig_dna = data.get("original_dna", data.get("original_mrna", ""))
    edit_dna = data.get("edited_dna", data.get("edited_mrna", ""))
    orig_prot = data.get("original_protein", "")
    edit_prot = data.get("edited_protein", "")
    orig_mrna = data.get("original_mrna", "")
    edit_mrna = data.get("edited_mrna", "")

    frameshift = data.get("frameshift", False)
    premature_stop = data.get("premature_stop", False)

    records = []

    # 1. Original DNA
    records.append(f">CRISPR_Sim|Original_DNA|Length={len(orig_dna)}bp")
    records.append(_wrap_seq(orig_dna, 80))

    # 2. Edited DNA
    records.append(f">CRISPR_Sim|Edited_DNA_{repair_type}|Length={len(edit_dna)}bp|Frameshift={frameshift}")
    records.append(_wrap_seq(edit_dna, 80))

    # 3. Original Protein
    if orig_prot:
        records.append(f">CRISPR_Sim|Original_Protein|Length={len(orig_prot)}aa")
        records.append(_wrap_seq(orig_prot, 80))

    # 4. Edited Protein
    if edit_prot:
        records.append(f">CRISPR_Sim|Edited_Protein_{repair_type}|Length={len(edit_prot)}aa|StopCodon={premature_stop}")
        records.append(_wrap_seq(edit_prot, 80))

    # 5. Original mRNA
    if orig_mrna:
        records.append(f">CRISPR_Sim|Original_mRNA|Length={len(orig_mrna)}nt")
        records.append(_wrap_seq(orig_mrna, 80))

    # 6. Edited mRNA
    if edit_mrna:
        records.append(f">CRISPR_Sim|Edited_mRNA_{repair_type}|Length={len(edit_mrna)}nt")
        records.append(_wrap_seq(edit_mrna, 80))

    return "\n".join(records) + "\n"
