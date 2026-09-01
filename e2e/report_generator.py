"""Generate E2E Excel report (PancreaScan-style layout)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class TestResult:
    test_id: str
    module: str
    name: str
    description: str
    steps: str
    expected: str
    actual: str
    status: str  # PASS | FAIL | SKIP
    severity: str
    duration_ms: float
    test_type: str
    executed_at: str
    remarks: str = ""


HEADER_FILL = PatternFill("solid", fgColor="006B76")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
SKIP_FILL = PatternFill("solid", fgColor="FFEB9C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="006B76")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _auto_width(ws, min_width=10, max_width=48):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row, col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def _style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def generate_report(
    results: list[TestResult],
    output_path: Path,
    *,
    app_name: str = "CRISPR-Sim",
    base_url: str,
    app_url: str | None = None,
    environment: str = "Production",
    tester: str = "Selenium E2E Automation",
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    total = len(results)
    passed = sum(1 for r in results if r.status == "PASS")
    failed = sum(1 for r in results if r.status == "FAIL")
    skipped = sum(1 for r in results if r.status == "SKIP")
    pass_rate = (passed / total * 100) if total else 0.0

    summary_rows = [
        ("E2E Test Execution Report", ""),
        ("Application", app_name),
        ("Report Generated", now),
        ("Environment", environment),
        ("API Base URL", base_url),
        ("App URL", app_url or "N/A"),
        ("Tester", tester),
        ("", ""),
        ("Total Test Cases", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Skipped", skipped),
        ("Pass Rate (%)", round(pass_rate, 2)),
        ("", ""),
        ("Module Breakdown", "Pass / Fail / Skip"),
    ]

    modules: dict[str, dict[str, int]] = {}
    for r in results:
        m = modules.setdefault(r.module, {"PASS": 0, "FAIL": 0, "SKIP": 0})
        m[r.status] = m.get(r.status, 0) + 1

    for mod, counts in sorted(modules.items()):
        summary_rows.append(
            (mod, f"{counts.get('PASS', 0)} / {counts.get('FAIL', 0)} / {counts.get('SKIP', 0)}")
        )

    ws_sum["A1"].font = TITLE_FONT
    for i, (k, v) in enumerate(summary_rows, start=1):
        ws_sum.cell(i, 1, k)
        ws_sum.cell(i, 2, v)
        ws_sum.cell(i, 1).font = Font(bold=True) if k and k not in ("E2E Test Execution Report", "") else Font()
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 55

    # ── Test Results sheet ────────────────────────────────────────────────
    ws = wb.create_sheet("Test Results")
    headers = [
        "S.No",
        "Test ID",
        "Module",
        "Test Case Name",
        "Description",
        "Test Steps",
        "Expected Result",
        "Actual Result",
        "Status",
        "Severity",
        "Type",
        "Duration (ms)",
        "Executed At",
        "Remarks",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    _style_header_row(ws, 1, len(headers))

    for i, r in enumerate(results, start=1):
        row = i + 1
        values = [
            i,
            r.test_id,
            r.module,
            r.name,
            r.description,
            r.steps,
            r.expected,
            r.actual,
            r.status,
            r.severity,
            r.test_type,
            round(r.duration_ms, 1),
            r.executed_at,
            r.remarks,
        ]
        fill = PASS_FILL if r.status == "PASS" else FAIL_FILL if r.status == "FAIL" else SKIP_FILL
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row, c, val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c == 9:
                cell.fill = fill
                cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"
    _auto_width(ws)

    # ── Failed Tests sheet ────────────────────────────────────────────────
    ws_fail = wb.create_sheet("Failed Tests")
    fail_headers = ["Test ID", "Module", "Test Case Name", "Expected", "Actual", "Remarks"]
    for c, h in enumerate(fail_headers, start=1):
        ws_fail.cell(1, c, h)
    _style_header_row(ws_fail, 1, len(fail_headers))
    fr = 2
    for r in results:
        if r.status == "FAIL":
            for c, val in enumerate(
                [r.test_id, r.module, r.name, r.expected, r.actual, r.remarks], start=1
            ):
                ws_fail.cell(fr, c, val).border = BORDER
            fr += 1
    if fr == 2:
        ws_fail.cell(2, 1, "No failures — all tests passed.")
    _auto_width(ws_fail)

    wb.save(output_path)
    return output_path
