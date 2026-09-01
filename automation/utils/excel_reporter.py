"""Enterprise Excel Report Generator using openpyxl."""

import os
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from automation.config.config import config
from automation.data.test_cases_catalog import TestCaseMetadata
from automation.utils.logger import logger


class ExcelReporter:
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or config.excel_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def _apply_header_style(self, ws, col_count: int):
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='medium', color='1F497D'),
        )
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 28

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    def generate_all_reports(self, test_results: List[TestCaseMetadata], metrics: Dict[str, Any]):
        """Generate all required Excel reports."""
        self._generate_master_report(test_results, metrics)
        self._generate_filtered_report(test_results, "PASSED", "Passed_Test_Cases.xlsx")
        self._generate_filtered_report(test_results, "FAILED", "Failed_Test_Cases.xlsx")
        self._generate_summary_report(test_results, metrics)
        logger.info(f"📊 Excel reports generated in: {self.output_dir}")

    def _generate_master_report(self, test_results: List[TestCaseMetadata], metrics: Dict[str, Any]):
        wb = Workbook()
        
        # Sheet 1: Executed Test Cases
        ws1 = wb.active
        ws1.title = "Executed Test Cases"
        headers1 = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Expected Result", "Actual Result"]
        ws1.append(headers1)
        self._apply_header_style(ws1, len(headers1))

        passed_tests = [t for t in test_results if t.status == "PASSED"]
        failed_tests = [t for t in test_results if t.status == "FAILED"]
        skipped_tests = [t for t in test_results if t.status in ("SKIPPED", "BLOCKED")]

        for row_idx, t in enumerate(test_results, start=2):
            ws1.append([
                t.test_id,
                t.module,
                f"{t.module} - {t.test_id}",
                t.status,
                round(t.execution_time_s, 3),
                t.priority,
                t.expected_result,
                t.actual_result or ("Passed without error" if t.status == "PASSED" else t.error_message),
            ])
            # Color status
            status_cell = ws1.cell(row=row_idx, column=4)
            if t.status == "PASSED":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
            elif t.status == "FAILED":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006", bold=True)
        self._auto_fit_columns(ws1)

        # Sheet 2: Passed Tests
        ws2 = wb.create_sheet(title="Passed Tests")
        ws2.append(["Test ID", "Module", "Priority", "Execution Time (s)", "Result"])
        self._apply_header_style(ws2, 5)
        for t in passed_tests:
            ws2.append([t.test_id, t.module, t.priority, round(t.execution_time_s, 3), "PASSED"])
        self._auto_fit_columns(ws2)

        # Sheet 3: Failed Tests
        ws3 = wb.create_sheet(title="Failed Tests")
        ws3.append(["Test ID", "Module", "Priority", "Error Message", "Screenshot", "Stack Trace"])
        self._apply_header_style(ws3, 6)
        for t in failed_tests:
            ws3.append([t.test_id, t.module, t.priority, t.error_message, t.screenshot_path or "N/A", t.stack_trace or "N/A"])
        self._auto_fit_columns(ws3)

        # Sheet 4: Skipped Tests
        ws4 = wb.create_sheet(title="Skipped Tests")
        ws4.append(["Test ID", "Module", "Priority", "Reason"])
        self._apply_header_style(ws4, 4)
        for t in skipped_tests:
            ws4.append([t.test_id, t.module, t.priority, t.error_message or "Precondition not met / skipped"])
        self._auto_fit_columns(ws4)

        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title="Execution Metrics")
        ws5.append(["Metric", "Value"])
        self._apply_header_style(ws5, 2)
        metrics_rows = [
            ("Deployment Base URL", metrics.get("base_url", config.base_url)),
            ("Total Test Cases", metrics.get("total_cases", len(test_results))),
            ("Passed Tests", len(passed_tests)),
            ("Failed Tests", len(failed_tests)),
            ("Skipped Tests", len(skipped_tests)),
            ("Pass Percentage", f"{metrics.get('pass_percent', 0.0):.2f}%"),
            ("Total Duration (s)", f"{metrics.get('total_duration_s', 0.0):.2f}"),
            ("Execution Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ]
        for m, v in metrics_rows:
            ws5.append([m, v])
        self._auto_fit_columns(ws5)

        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title="Defect Summary")
        ws6.append(["Defect ID", "Module", "Severity", "Root Cause", "Failed Step"])
        self._apply_header_style(ws6, 5)
        for idx, t in enumerate(failed_tests, start=1):
            ws6.append([f"DEF-{idx:03d}", t.module, t.priority, t.error_message or "Assertion Failure", t.steps])
        self._auto_fit_columns(ws6)

        filepath = os.path.join(self.output_dir, "Automation_Test_Report.xlsx")
        wb.save(filepath)

    def _generate_filtered_report(self, test_results: List[TestCaseMetadata], status_filter: str, filename: str):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{status_filter.capitalize()} Cases"
        headers = ["Test ID", "Module", "Priority", "Execution Time (s)", "Details"]
        ws.append(headers)
        self._apply_header_style(ws, len(headers))
        
        filtered = [t for t in test_results if t.status == status_filter]
        for t in filtered:
            detail = t.error_message if status_filter == "FAILED" else "Execution Successful"
            ws.append([t.test_id, t.module, t.priority, round(t.execution_time_s, 3), detail])
        self._auto_fit_columns(ws)
        wb.save(os.path.join(self.output_dir, filename))

    def _generate_summary_report(self, test_results: List[TestCaseMetadata], metrics: Dict[str, Any]):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        headers = ["Module", "Total Cases", "Passed", "Failed", "Pass Rate (%)"]
        ws.append(headers)
        self._apply_header_style(ws, len(headers))

        modules = sorted(list(set(t.module for t in test_results)))
        for m in modules:
            m_tests = [t for t in test_results if t.module == m]
            m_pass = len([t for t in m_tests if t.status == "PASSED"])
            m_fail = len([t for t in m_tests if t.status == "FAILED"])
            rate = (m_pass / len(m_tests)) * 100 if m_tests else 0
            ws.append([m, len(m_tests), m_pass, m_fail, f"{rate:.1f}%"])
        self._auto_fit_columns(ws)
        wb.save(os.path.join(self.output_dir, "Summary_Report.xlsx"))
